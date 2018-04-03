from lib.logger import StreamFileLogger
import openpyxl
from openpyxl.styles import colors,PatternFill
from openpyxl import Workbook
from lib.excel import Excel
from utils import settings

def read_src_excel(sheetname):
    excel = Excel(settings.SRC_FILE_PATH)
    srccolnames = excel.get_column_names(sheetname)
    src_colnames = []
    for src_col_item in srccolnames:
        if src_col_item is not None:
            src_colnames.append(src_col_item)
    return src_colnames

def read_tgt_excel(sheetname):
    excel = Excel(settings.TGT_FILE_PATH)
    tgtcolnames = excel.get_column_names(sheetname)
    tgt_colnames = []
    for tgt_col_item in tgtcolnames:
        if tgt_col_item is not None:
            tgt_colnames.append(tgt_col_item)
    return tgt_colnames

def get_dif_columns(sheetname):
    src_items = read_src_excel(sheetname)
    tgt_items = read_tgt_excel(sheetname)
    dif_items = []

    for item in tgt_items:
        if item not in src_items:
            dif_items.append(item)

    #print(dif_items)
    return dif_items

def highlight_columns(sheetname):
    dif_items = get_dif_columns(sheetname)
    excel = Excel(settings.TGT_FILE_PATH)
    wb = excel.get_wb()
    ws = excel.get_sheet(sheetname)
    max_row = excel.get_max_row(sheetname)
    col_hdr = []
    for dif_col in dif_items:
        col_hdr_item = excel.convert_col2header(sheetname=sheetname,column_name=dif_col)
        col_hdr.append(col_hdr_item)

    #print(col_hdr)

    # for row_nbr in range(1,max_row-490):
    #     for hdr in col_hdr:
    #         cell_item = hdr+str(row_nbr)
    #         print(cell_item)
    #         cell_item.fill = openpyxl.styles.fills.Color.
    #         ws[cell_item].fill = colors.RED
    #
    # wb.save('new_compare_result.xlsx')

    for item in get_dif_columns(sheetname):
        cells = excel.get_column(sheetname,column_name=item)
        print(cells)
        for cell in cells:
            cell.fill=PatternFill(patternType='solid', fgColor='00FF0000')

    wb.save('new_compare_result.xlsx')

#highlight_columns('NKTP')




'''
1. read excel v1
2. get V1 column_name
3. loop through get V1 data into message {{name:A,age:B,grade:C},{count:0}}
4. create vv2 based on v2
5. get VV2 column_names
6. compare VV2 column_names with v1 col_names
7. If vv2 col_names is different from v1 col_names find the difference
8. loop through vv2 get data into message (exclude difference columns)
9. compare vv2 message with v1 message
10. mark the differences

'''
