import views.delete.getMessageData as getMsgData
import views.delete.getCompareColName as getColumn
from etc import settings
from lib.excel import Excel
from openpyxl.styles import PatternFill
from lib.logger import StreamFileLogger

_sflogger = StreamFileLogger(settings.LOG_FILE, __file__).get_logger()
def get_diff_rowdNum(srcTuple,tgtTuple):
    _srcData = srcTuple[0]
    _tgtData = tgtTuple[0]
    _tgtStartNum = tgtTuple[1]
    _numlist = []
    lineNum = _tgtStartNum
    for _tgtitem in _tgtData:
        if (_tgtitem not in _srcData):
            _numlist.append(lineNum)
        lineNum += 1
    return _numlist

def get_matchIdx_rowdNum(srcTuple,tgtTuple):
    _srcData = srcTuple[0]
    _srcStartNum = srcTuple[1]

    _tgtData = tgtTuple[0]
    _tgtStartNum = tgtTuple[1]
    #store match rowNum in both file
    _numlist = []
    for _i in range(0,len(_srcData)-1):
        for _j in range(0,len(_tgtData)-1):
            if _srcData[_i] == _tgtData[_j]:
                list1 = (str(_i+_srcStartNum)+','+str(_j+_tgtStartNum)).split(',')
                _numlist.append(list1)
                break
    return _numlist


def setBgColorRow(srcexcel,tgtexcel,sheetname):
    _sflogger.info('Compare start:')
    _match_column_name = getColumn.get_match_columns(srcexcel, tgtexcel, sheetname)
    _srcTuple = getMsgData.get_srcdata_message(srcexcel, _match_column_name, sheetname)
    _tgtTuple = getMsgData.get_tgtdata_message(tgtexcel, _match_column_name, sheetname)
    _sflogger.info('Start get different row number:')
    _getrowsNum = get_diff_rowdNum(_srcTuple,_tgtTuple)
    _wb = tgtexcel.get_wb()
    _ws = tgtexcel.get_sheet(sheetname)
    _sflogger.info('Start highlight different row color:')
    for curitem in _ws.iter_rows():

        if curitem[0].row in _getrowsNum:
            for cell in curitem:
                cell.fill = PatternFill(fgColor = 'EE7600', fill_type = 'solid')

    _sflogger.info('Start highlight new added column color:')
    _addColumn = getColumn.get_add_columns(srcexcel, tgtexcel, sheetname)
    if _addColumn is not None:
        #convert add column name into excel head(A B C D AA...)
        for i in range(0, len(_addColumn)):
            _addColumn[i]  = getColumn.conver_header(_ws, _addColumn[i])

        for _row in _ws.iter_rows():
            for _cellitem in _row:
                if _cellitem.column in _addColumn:
                    _cellitem.fill = PatternFill(fgColor='87CEEB', fill_type='solid')



    _sflogger.info('Finish comparision:')
    try:
        _wb.save(settings.END_FILE_PATH)
        _sflogger.info('Save completed')
    except Exception as e:
        _sflogger.error('Execute failed.', exc_info=True)

def setBgColorRowIdx(srcexcel,tgtexcel,sheetname,idx):
    _sflogger.info('Compare start:')
    #initial getmessage data
    _match_column_name = getColumn.get_match_columns(srcexcel,tgtexcel,sheetname,idx)
    _srcTuple = getMsgData.get_srcdata_message(srcexcel, _match_column_name, sheetname, idx)
    _tgtTuple = getMsgData.get_tgtdata_message(tgtexcel, _match_column_name, sheetname, idx)

    _getrowsNum = get_matchIdx_rowdNum(_srcTuple,_tgtTuple)
    _wb = tgtexcel.get_wb()
    _ws = tgtexcel.get_sheet(sheetname)
    # only flag target file
    _srcws = srcexcel.get_sheet(sheetname)

    _getZips= getMsgData.get_compare_colNum(srcexcel,tgtexcel,sheetname,idx)
    _sflogger.info('Start highlight updated data :')
    #set color in same index but different cell value
    for _row in _getrowsNum:
        for _zip in _getZips:
            #getbothCellsName _zip[0] is srcrownum,_zip[1] is tgrrownum.row is same
            _srccellname = "{}{}".format(_zip[0], _row[0])
            _tgtcellname = "{}{}".format(_zip[1], _row[1])
            _srclvalue = _srcws[_srccellname].value
            _tgtlvalue = _ws[_tgtcellname].value
            _srclvalue = str(_srclvalue).strip().upper()
            _tgtlvalue = str(_tgtlvalue).strip().upper()
            if(_srclvalue !=_tgtlvalue):
                _ws[_tgtcellname].fill = PatternFill(fgColor = 'EE7600', fill_type = 'solid')

    _sflogger.info('Start highlight new added row data :')
    _getdiffrowsNum = get_diff_rowdNum(_srcTuple,_tgtTuple)
    for curitem in _ws.iter_rows():

        if curitem[0].row  in _getdiffrowsNum:
            for cell in curitem:
                cell.fill = PatternFill(fgColor = 'EEC900', fill_type = 'solid')

    _sflogger.info('Start highlight new added column data :')
    _addColumn = getColumn.get_add_columns(srcexcel, tgtexcel, sheetname)
    if _addColumn is not None:
        # convert add column name into excel head(A B C D AA...)
        for i in range(0, len(_addColumn)):
            _addColumn[i] = getColumn.conver_header(_ws, _addColumn[i])

        for _row in _ws.iter_rows():
            for _cellitem in _row:
                if _cellitem.column in _addColumn:
                    _cellitem.fill = PatternFill(fgColor='87CEEB', fill_type='solid')
    _sflogger.info('Finished comparison')
    try:
        _wb.save(settings.END_FILE_PATH)
        _sflogger.info('Save completed')
    except Exception as e:
        _sflogger.error('Execute failed.', exc_info=True)




#
# def test():
#     _srcpath = settings.SRC_FILE_PATH
#     _tgtpath = settings.TGT_FILE_PATH
#     _srcexcel = Excel(_srcpath)
#     _tgtexcel = Excel(_tgtpath)
#     #setBgColorRowIdx(_srcexcel,_tgtexcel,'CAPS Industry KPIs New','PRIMARY CONTACT_EMAIL')
#     setBgColorRow(_srcexcel, _tgtexcel, 'CAPS Industry KPIs New')
#
# test()

