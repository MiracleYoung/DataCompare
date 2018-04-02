import views.delete.getMessageData as getMsgData
from utils import settings
from lib.excel import Excel
from lib.logger import StreamFileLogger
from openpyxl.styles import PatternFill

def get_diff_rowdata(sheetname):
    _srcpath = settings.SRC_FILE_PATH
    _tgtpath = settings.TGT_FILE_PATH
    _srcData = getMsgData.get_data_message(_srcpath,sheetname)
    _tgtData = getMsgData.get_data_message(_tgtpath,sheetname)
    _numlist = []
    lineNum = 2
    for _tgtitem in _tgtData:
        if (_tgtitem not in _srcData):
            _numlist.append(lineNum)
        lineNum += 1
    return _numlist


def setBgColor(sheetname):
    _getSetList = get_diff_rowdata(sheetname)
    _filepath = settings.END_FILE_PATH
    _excel = Excel(_filepath)
    _ws = _excel.get_sheet(sheetname)
    for curitem in _ws.iter_rows():

        if curitem[0].row in _getSetList:

            print(curitem[0].row)
            for cell in curitem:
                cell.fill = PatternFill(fgColor = 'FF0000', fill_type = 'solid')


setBgColor('CAPS Industry KPIs New')
