#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 3/14/18 3:48 PM
# @Author  : Miracle Young
# @File    : excel.py

from openpyxl import load_workbook

from lib.logger import StreamFileLogger
from etc import settings

_sflogger = StreamFileLogger(settings.LOG_FILE, __file__).get_logger()


class Excel:
    def __init__(self, path):
        self._wb = load_workbook(filename=path, data_only=True)
        _sflogger.info('Open excel: {}'.format(path))

    def get_wb(self):
        return self._wb

    def get_sheet(self, sheetname):
        return self._wb[sheetname]

    def get_all_sheetnames(self):
        return self._wb.get_sheet_names()

    def get_max_row(self, sheetname):
        return self.get_sheet(sheetname).max_row

    def get_max_column(self, sheetname):
        return self.get_sheet(sheetname).max_column

    def get_dimensions(self, sheetname):
        return self.get_sheet(sheetname).dimensions.split(':')

    def get_column_names(self, sheetname, start=None, end=None, mapping=None):
        [_start, _end] = [start, end] if start and end else self.get_dimensions(sheetname)
        _sheet = self.get_sheet(sheetname)
        if mapping:
            _columns = [mapping.get(_column.value) for _i, _column in enumerate(_sheet.rows) if mapping]
        else:
            _columns = [_column.value for _column in _sheet[_start:_end][0]]
        return _columns

    def convert_col2header(self, sheetname, column_name):
        _sheet = self.get_sheet(sheetname)
        for _row in _sheet:
            for _j, _cell in enumerate(_row):
                if _cell.value.upper() == column_name.upper():
                    return _cell.column
        return ''

    def get_column(self, sheetname, column_name):
        _sheet = self.get_sheet(sheetname)
        _header = self.convert_col2header(sheetname, column_name)
        yield from _sheet[_header] if _header else []

    def read_excel_by_row(self, sheetname, start=None, end=None, mapping=None):
        try:
            [_start, _end] = [start, end] if start and end else self.get_dimensions(sheetname)
            _sheet = self.get_sheet(sheetname)
            _sflogger.debug('Read {}{}{}: starting...'.format(sheetname, _start, _end))
            _columns = self.get_column_names(sheetname, _start, _end, mapping)
            _sflogger.debug('Columns: {}'.format(_columns))
            yield from _sheet[_start: _end][1:]
        except Exception as e:
            _sflogger.error('Execute failed.', exc_info=True)
            return []
